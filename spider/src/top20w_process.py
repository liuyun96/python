# -*- coding:utf-8 -*
'''
淘宝TOP20W的Excel数据处理程序，能够自动分析Excel文件，并将其中的词语作为paoding的淘宝专用词库进行存储。
使用方法为 python top20w_process.py <excel file>
Created on 2012-03-16
r'c:/top20w2_15.xlsx'
@author: lbj
'''
from openpyxl.reader.excel import load_workbook  
# 导入mysqldb，用于连接mysql数据库
import MySQLdb
# 导入正则表达式。
import re
import sys
import datetime

from Excel import Excel2003;

def extractDigitalFromInfo(info):
    m = digitalPattern.findall(info)
    if(len(m) > 0):
        return m[len(m) - 1];
    print info + '提取数字失败';
    return 0;
if len(sys.argv) != 2:
    print '使用方法:python top20w_process.py <excel file>'
    sys.exit(0)
conn = MySQLdb.connect(host='192.168.0.245', user='myuser', passwd='myuser', charset="utf8", db="dongji", use_unicode=True)
conn.autocommit(True)
# 提取数字的正则表达式。
global digitalPattern
digitalPattern = re.compile('([.\d]+)')
cursor = conn.cursor()
sql = "DELETE FROM TOP20W"
cursor.execute(sql)
#sql = "DELETE FROM TOP_TB_KEYWORD"
#cursor.execute(sql)

fileName = sys.argv[1];
if fileName.find('xlsx') !=-1:
    # 读取excel2007文件  
    wb = load_workbook(filename=sys.argv[1], use_iterators=True)  
    # 取第一张表  
    sheetnames = wb.get_sheet_names()  
    for  sname in sheetnames:
        if sname == "首页":
            continue;
        print sname;
        ws = wb.get_sheet_by_name(sname)
        # 读入现有词库,作为set对象。
        # 显示表名，表行数，表列数
        datas = [];
        line = 0  
        for row in ws.iter_rows():  # it brings a new method: iter_rows()
            line = line + 1
            if(line <= 1):
                continue
            keyword = row[0].internal_value
            yiCate = row[1].internal_value
            erCate = row[2].internal_value
            keywordBuyerInfo = str(row[3].internal_value);#关键词
            keywordClickInfo = str(row[4].internal_value);#点击量
            keywordPPCInfo = str(row[5].internal_value);#ppc
            keywordInfo = keywordBuyerInfo + "," + keywordClickInfo + "," + keywordPPCInfo;
            keywordBuyerCount = extractDigitalFromInfo(keywordBuyerInfo); 
            keywordClickCount = extractDigitalFromInfo(keywordClickInfo); 
            keywordPPCPrice = extractDigitalFromInfo(keywordPPCInfo); 
            if(type(keyword) is float):
                keyword = str(int(keyword));
            elif(type(keyword) is datetime.time):
                keyword = str(keyword.hour) + ':' + str(keyword.minute)
            datas.append((yiCate,erCate, keyword,
                keywordBuyerCount, keywordClickCount, keywordPPCPrice, keywordInfo));
            if(line % 50000 == 0):
                print line
                cursor.executemany('''INSERT INTO TOP20W (FIRST_CLASS,SECOND_CLASS,KEYWORD,KEYWORD_BUYER_COUNT,CLICK_COUNT,PPC_PRICE,KEYWORD_INFO,PUB_DATE)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,curdate())''', tuple(datas));
                del datas[:]
        cursor.executemany('''INSERT INTO TOP20W (FIRST_CLASS,SECOND_CLASS,KEYWORD,KEYWORD_BUYER_COUNT,CLICK_COUNT,PPC_PRICE,KEYWORD_INFO,PUB_DATE)
            VALUES(%s,%s,%s,%s,%s,%s,%s,sysdate())''', tuple(datas));
else:
    #读取2003
    #path = '/home/liuy/download/top20w1104.xls'; 
    excel = Excel2003(fileName);
    for j in range(1,7):
        excel.setSheet(j);
        datas = [];
        line = 0
        for i in range(1, excel.Y):
            line = line + 1;
            keyword = excel.read(i,0);
            yiCate = excel.read(i,1);
            erCate = excel.read(i,2);
            sanCate = str(excel.read(i,3)).replace('0','');
            keywordBuyerInfo = str(excel.read(i,4));#关键词购买人数
            keywordClickInfo = str(excel.read(i,5));#点击量
            keywordPPCInfo = str(excel.read(i,6));#ppc
            keywordInfo = keywordBuyerInfo + "," + keywordClickInfo + "," + keywordPPCInfo;
            keywordBuyerCount = extractDigitalFromInfo(keywordBuyerInfo); 
            keywordClickCount = extractDigitalFromInfo(keywordClickInfo); 
            keywordPPCPrice = extractDigitalFromInfo(keywordPPCInfo); 
            
            if(type(keyword) is float):
                keyword = str(int(keyword));
            elif(type(keyword) is datetime.time):
                keyword = str(keyword.hour) + ':' + str(keyword.minute)
            datas.append((yiCate,erCate,sanCate, keyword,
                keywordBuyerCount, keywordClickCount, keywordPPCPrice, keywordInfo));
            if(line % 10000 == 0):
                print line
                cursor.executemany('''INSERT INTO TOP20W (FIRST_CLASS,SECOND_CLASS,THIRD_CLASS,KEYWORD,KEYWORD_BUYER_COUNT,CLICK_COUNT,PPC_PRICE,KEYWORD_INFO,PUB_DATE)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,curdate())''', tuple(datas));
                del datas[:]
        cursor.executemany('''INSERT INTO TOP20W (FIRST_CLASS,SECOND_CLASS,THIRD_CLASS,KEYWORD,KEYWORD_BUYER_COUNT,CLICK_COUNT,PPC_PRICE,KEYWORD_INFO,PUB_DATE)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,curdate())''', tuple(datas));


