#-*- coding:utf-8 -*
'''
中国移动WAP网关IP地址Excel处理程序。
获取省份和网关IP地址列表
使用方法为 python wapip_process.py <excel file>
Created on 2012-03-16
r'c:/top20w2_15.xlsx'
@author: lbj
'''
from openpyxl.reader.excel import load_workbook  
#导入文件编码器，用于解析和生成utf8文件。
import sys
import MySQLdb

#导入正则表达式。
import re
def extractIPFromInfo(info):
    if(info==None):
        return "";
    m = IPPattern.findall(info)
    if(len(m)>0):
        return m[len(m)-1];
    print info+'提取ip失败';
    return "";
conn = MySQLdb.connect(host='127.0.0.1', user='root',passwd='hacker',charset = "utf8",db="viva_wap", use_unicode = True)  
conn.autocommit(True)
    
#提取数字的正则表达式。
IPPattern = re.compile('([.\d\./]+)')
iprange = re.compile('(\d+\.\d+\.\d+\.)(\d+)/(\d+)')
line = 0
m = iprange.findall("221.179.36.17/28")
if(len(m)>0):
    print m[0][0],m[0][1],m[0][2]
#读取excel2007文件  
wb = load_workbook(filename = 'd:\\wapip.xlsx', use_iterators = True)  
#取第二张表  
sheetnames = wb.get_sheet_names()  
ws = wb.get_sheet_by_name(sheetnames[1])

cursor = conn.cursor()
for row in ws.iter_rows(): # it brings a new method: iter_rows()
    try:
        line=line+1
        if(line<=2):
            continue
        province =  row[1].internal_value;
        if(province==None):
            break;
        ip1 = extractIPFromInfo(row[8].internal_value);
        ip2 = extractIPFromInfo(row[9].internal_value);
        if(len(ip1)==0):
            ip1 = ip2;
        if(len(ip1)==0 and len(ip2)==0):
            print province,ip1,ip2
        m = iprange.findall(ip1)
        if(len(m)>0):
            for i in range(int(m[0][1]),int(m[0][1])+int(m[0][2])):
                ip = m[0][0]+str(i)  
                print ip
                cursor.execute('''INSERT INTO WAP_GATEWAY_IP (IP,PROVINCE) VALUES(%s,%s)''',(ip,province));
        else:      
            cursor.execute('''INSERT INTO WAP_GATEWAY_IP (IP,PROVINCE) VALUES(%s,%s)''',(ip1,province));
    except:
        print sys.exc_info() #print all traceback exceptions, for debugging    
        print province,ip1,ip2
